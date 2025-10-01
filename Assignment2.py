import psycopg2
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

os.makedirs("charts", exist_ok=True)
os.makedirs("exports", exist_ok=True)

conn = psycopg2.connect(
    dbname='postgres',
    user='postgres',
    password='732Alshyn',
    host='localhost',
    port=5433
)

dataframes_dict = {}

def save_and_report(df, filename, chart_type, description):
    """Helper function to save chart and print console report."""
    print(f"{chart_type}: {len(df)} rows  saved to {filename}. {description}")


query_pie = """
SELECT COUNT(c.customer_id) AS number_of_clients, op.payment_type
FROM olist_customers_dataset AS c 
JOIN olist_orders_dataset AS o ON c.customer_id = o.customer_id 
JOIN olist_order_payments_dataset AS op ON o.order_id = op.order_id
GROUP BY op.payment_type
"""
df_pie = pd.read_sql(query_pie, conn)
plt.figure(figsize=(8, 6))
plt.pie(df_pie['number_of_clients'],
        labels=df_pie['payment_type'],
        autopct='%1.1f%%', startangle=90)
plt.title("Distribution of Clients by Payment Type")
plt.savefig("charts/pie_payment_type.png")
plt.close()
save_and_report(df_pie, "charts/pie_payment_type.png", "Pie Chart",
                "Shows distribution of clients by payment type.")
dataframes_dict["Clients by Payment Type"] = df_pie

query_bar = """
SELECT c.customer_city, COUNT(DISTINCT o.customer_id) AS customers_with_orders
FROM olist_orders_dataset o
JOIN olist_customers_dataset c ON o.customer_id = c.customer_id
JOIN olist_order_items_dataset oi ON o.order_id = oi.order_id
GROUP BY c.customer_city
ORDER BY customers_with_orders DESC
LIMIT 10;
"""
df_bar = pd.read_sql(query_bar, conn)
plt.figure(figsize=(10, 6))
plt.bar(df_bar['customer_city'], df_bar['customers_with_orders'], color="skyblue")
plt.title("Top 10 Cities by Number of Customers with Orders")
plt.xlabel("City")
plt.ylabel("Customers with Orders")
plt.xticks(rotation=45)
plt.savefig("charts/bar_customers_city_orders.png")
plt.close()
save_and_report(df_bar, "charts/bar_customers_city_orders.png", "Bar Chart",
                "Top 10 cities ranked by customers who placed orders.")
dataframes_dict["Top Cities Customers"] = df_bar

query_barh = """
SELECT s.seller_id, s.seller_city, ROUND(SUM(oi.price)::numeric,2) AS revenue
FROM olist_order_items_dataset oi
JOIN olist_sellers_dataset s ON oi.seller_id = s.seller_id
JOIN olist_orders_dataset o ON oi.order_id = o.order_id
GROUP BY s.seller_id, s.seller_city
ORDER BY revenue DESC
LIMIT 10;
"""
df_barh = pd.read_sql(query_barh, conn)
plt.figure(figsize=(10, 6))
plt.barh(df_barh['seller_id'], df_barh['revenue'], color="lightgreen")
plt.title("Top 10 Sellers by Revenue")
plt.xlabel("Revenue")
plt.ylabel("Seller ID")
plt.savefig("charts/barh_top_sellers_revenue.png")
plt.close()
save_and_report(df_barh, "charts/barh_top_sellers_revenue.png", "Horizontal Bar Chart",
                "Top 10 sellers ranked by revenue.")
dataframes_dict["Top Sellers Revenue"] = df_barh

query_line = """
SELECT to_char(DATE_TRUNC('month', o.order_purchase_timestamp), 'YYYY-MM') AS month,
       ROUND(SUM(oi.price)::numeric,2) AS total_sales
FROM olist_orders_dataset o
JOIN olist_order_items_dataset oi ON o.order_id = oi.order_id
JOIN olist_products_dataset p ON oi.product_id = p.product_id
GROUP BY month
ORDER BY month;
"""
df_line = pd.read_sql(query_line, conn)
plt.figure(figsize=(10, 6))
plt.plot(df_line['month'], df_line['total_sales'], marker="o", linestyle="-", color="orange")
plt.xticks(rotation=45)
plt.title("Monthly Total Sales")
plt.xlabel("Month")
plt.ylabel("Total Sales (R$)")
plt.savefig("charts/line_monthly_sales.png")
plt.close()
save_and_report(df_line, "charts/line_monthly_sales.png", "Line Chart",
                "Total monthly sales trend.")
dataframes_dict["Monthly Sales"] = df_line

query_hist = """
SELECT p.product_description_length
FROM olist_order_items_dataset oi
JOIN olist_products_dataset p ON oi.product_id = p.product_id
JOIN olist_orders_dataset o ON oi.order_id = o.order_id
WHERE p.product_description_length IS NOT NULL;
"""
df_hist = pd.read_sql(query_hist, conn)

plt.figure(figsize=(8, 6))
plt.hist(df_hist['product_description_length'], bins=30, color="purple", edgecolor="black")
plt.title("Distribution of Product Description Length (Sold Products)")
plt.xlabel("Description Length")
plt.ylabel("Frequency")
plt.savefig("charts/hist_product_description_length.png")
plt.close()



query_scatter = """
SELECT oi.price, oi.freight_value, p.product_category_name
FROM olist_order_items_dataset oi
JOIN olist_products_dataset p ON oi.product_id = p.product_id
JOIN olist_orders_dataset o ON oi.order_id = o.order_id
WHERE oi.price IS NOT NULL AND oi.freight_value IS NOT NULL
LIMIT 500;
"""
df_scatter = pd.read_sql(query_scatter, conn)
plt.figure(figsize=(8, 6))
plt.scatter(df_scatter['price'], df_scatter['freight_value'],
            c=pd.factorize(df_scatter['product_category_name'])[0],
            cmap='tab10', alpha=0.6)
plt.title("Price vs Freight Value by Product Category")
plt.xlabel("Price")
plt.ylabel("Freight Value")
plt.savefig("charts/scatter_price_vs_freight.png")
plt.close()
save_and_report(df_scatter, "charts/scatter_price_vs_freight.png", "Scatter Plot",
                "Price vs freight value colored by product category.")
dataframes_dict["Price vs Freight"] = df_scatter

query_plotly = """
SELECT to_char(DATE_TRUNC('month', o.order_purchase_timestamp), 'YYYY-MM') AS month,
       p.product_category_name,
       ROUND(SUM(oi.price)::numeric, 2) AS sales
FROM olist_orders_dataset o
JOIN olist_order_items_dataset oi ON o.order_id = oi.order_id
JOIN olist_products_dataset p ON oi.product_id = p.product_id
GROUP BY month, p.product_category_name
ORDER BY month;
"""
df_plotly = pd.read_sql(query_plotly, conn)

top_categories = df_plotly.groupby("product_category_name")["sales"].sum().nlargest(5).index
df_top = df_plotly[df_plotly["product_category_name"].isin(top_categories)]

fig = px.line(
    df_top,
    x="month",
    y="sales",
    color="product_category_name",
    title="Monthly Sales by Top 5 Product Categories",
    markers=True
)

fig.update_layout(
    xaxis_title="Month",
    yaxis_title="Sales (R$)",
    legend_title="Category",
    xaxis=dict(
        rangeslider=dict(visible=True),
    )
)

fig.show()
dataframes_dict["Monthly Sales by Category"] = df_top

def export_to_excel(dataframes_dict, filename):
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        for sheet_name, df in dataframes_dict.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)

    wb = load_workbook(filename)
    for ws in wb.worksheets:
        ws.freeze_panes = "B2" 
        ws.auto_filter.ref = ws.dimensions

        for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
            if all(isinstance(cell.value, (int, float)) for cell in col if cell.value is not None):
                col_letter = col[0].column_letter
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    ColorScaleRule(start_type="min", start_color="FFAA0000",
                                   mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                                   end_type="max", end_color="FF00AA00")
                )
    wb.save(filename)
    print(f"Created file {filename}, {len(dataframes_dict)} sheets, "
          f"{sum(len(df) for df in dataframes_dict.values())} rows.")

export_to_excel(dataframes_dict, "exports/report.xlsx")

conn.close()